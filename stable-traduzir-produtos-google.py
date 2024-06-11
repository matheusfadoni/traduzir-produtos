import os
import pandas as pd
from googletrans import Translator
from marcas import marcas
from openpyxl import load_workbook
from openpyxl import Workbook

# Carregar o arquivo XLSX
nome_arquivo = 'Parte5'
file_path = nome_arquivo+'.xlsx'
df = pd.read_excel(file_path, engine='openpyxl')

# Transformar todas as strings em letras maiúsculas
df = df.apply(lambda x: x.map(lambda x: x.upper() if isinstance(x, str) else x))

# Inicializar o tradutor
translator = Translator()

# Função para traduzir mantendo as regras especificadas
def traduzir(text):
    if isinstance(text, str):
        try:
            traduzido = translator.translate(text, src='es', dest='pt').text
            print(f"Traduzido '{text}' para '{traduzido}'")
            return traduzido.upper()  # Transformar tradução em maiúsculas
        except Exception as e:
            print(f"Erro ao traduzir o texto '{text}': {e}")
            return text
    return text

# Nome do arquivo de saída
output_file = nome_arquivo+'-revisar.xlsx'
# Verificar se o arquivo de saída já existe
if not os.path.exists(output_file):
    # Criar um novo workbook
    workbook = Workbook()
    # Adicionar uma planilha ativa
    worksheet = workbook.active
    # Salvar o workbook
    workbook.save(output_file)

# Carregar o workbook existente
workbook = load_workbook(output_file)
worksheet = workbook.active

# Itera sobre as linhas do DataFrame
for index, row in df.iterrows():
    text_to_translate = row['esp']
    
    # Verificar se o texto é uma string
    if isinstance(text_to_translate, str):
        marca_temporario = []
        # Substituir marcas no texto original por um marcador temporário
        for marca in marcas:
            if ' '+marca+' ' in text_to_translate:
                marca_temporario.append(' '+marca)
                text_to_translate = text_to_translate.replace(marca, ' ¿ ')
        translated_text = traduzir(text_to_translate)
        # Restaurar as marcas no texto traduzido
        if marca_temporario != []:
            print('___________________________')
            #marca_temporario.reverse()
            print(marca_temporario)
            print('___________________________')
        for marca in marca_temporario:
            if '>' in translated_text:
                translated_text = translated_text.replace('¿', marca, 1)
        row['trad'] = translated_text
    else:
        # Se não for string, manter o valor original
        row['trad'] = text_to_translate
    
    worksheet.append(row.tolist())
    print(f"Linha {index + 1} traduzida: {row['trad']}")

# Salvar o workbook
workbook.save(output_file)
print("Tradução completa!")