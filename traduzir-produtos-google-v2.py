import os
import re
import pandas as pd
from googletrans import Translator
from marcas import marcas
from openpyxl import load_workbook, Workbook
import nltk
from nltk.corpus import words as nltk_words

import colorama
from colorama import Fore

# Baixar dicionário de palavras da NLTK
nltk.download('words')

# Lista de palavras em inglês
english_words = set(nltk_words.words())
custom_words_to_add = {"WHISKY","whisky","POUNDS","pounds","nuts","NUTS","ECOTANK","ecotank","RYZEN","ryzen","CELULAR","celular"}
english_words.update(custom_words_to_add)
custom_words_to_remove = {"COLOR","PANTALON","PROTECTOR","LIQUOR","LECTOR","REFLECTOR","TAPA","MANGA","CORTA","DYE","INOXIDABLE","EL","DORADO","HACIENDA","BOTELLA","JARRA","CABEL","PELOTA","AUTO","TRILLO","EN","FUERTE","CASCO","MARRON","AMARILLO","AURICULAR","NEGRO", "BLANCO", "VINO", "GRIS", "CAJA", "RON", "CALZA", "AIRE", "GORRA", "CON", "COPA", "DINERO", "AA", "A", "P", "DE", "ES", "SH", "ERE","Y","SIN","MANO"}
custom_words_to_remove_lower = {"color","pantalon","protector","liquor","lector","reflector","tapa","manga","corta","dye","inoxidable","el","dorado","hacienda","botella","jarra","cable","pelota","auto","trillo","en","fuerte","casco","marron","amarillo","auricular","negro", "blanco", "vino", "gris", "caja", "ron", "calza", "aire", "gorra", "con", "copa", "dinero", "aa", "a", "p", "de", "es", "sh", "ere","y","sin","mano"}
english_words.difference_update(custom_words_to_remove)
english_words.difference_update(custom_words_to_remove_lower)

# Carregar o arquivo XLSX
nome_arquivo = 'Parte12'
file_path = nome_arquivo + '.xlsx'
df = pd.read_excel(file_path, engine='openpyxl')

# Transformar todas as strings em letras maiúsculas
df = df.apply(lambda x: x.map(lambda x: x.upper() if isinstance(x, str) else x))

# Inicializar o tradutor
translator = Translator()

# Função para identificar termos em inglês
def find_english_terms(text):
    terms = re.findall(r'\b[A-Za-z]+\b', text)
    english_terms = [term for term in terms if term.lower() in english_words]
    return english_terms

# Função para substituir termos em inglês por marcadores temporários
def replace_english_terms(text):
    english_terms = find_english_terms(text)
    terms_to_replace = {}
    for i, term in enumerate(english_terms):
        marker = f"(({i}))"
        text = re.sub(rf'\b{re.escape(term)}\b', marker, text)
        terms_to_replace[marker] = term
    return text, terms_to_replace

# Função para restaurar os termos em inglês
def restore_english_terms(text, terms_to_replace):
    for marker, term in terms_to_replace.items():
        text = text.replace(marker, term)
    return text

# Função para traduzir mantendo as regras especificadas
def traduzir(text):
    if isinstance(text, str):
        try:
            # Substituir termos em inglês por marcadores temporários
            text, terms_to_replace = replace_english_terms(text)
            print(terms_to_replace)
            # Traduzir o texto
            traduzido = translator.translate(text, src='es', dest='pt').text
            # Restaurar os termos em inglês
            traduzido = restore_english_terms(traduzido, terms_to_replace)
            traduzido = clean_string(traduzido)
            print(f"Traduzido '{text}' para '{traduzido}'")
            return traduzido.upper()  # Transformar tradução em maiúsculas
        except Exception as e:
            print(f"Erro ao traduzir o texto '{text}': {e}")
            return text
    return text

def clean_string(text):
    # Remove parênteses '(' e ')'
    text = text.replace('(', '')
    text = text.replace(')', '')
    # Substitui múltiplos espaços por um único espaço
    text = re.sub(r'\s+', ' ', text)
    return text

# Nome do arquivo de saída
output_file = nome_arquivo + '-revisar.xlsx'
# Verificar se o arquivo de saída já existe
if not os.path.exists(output_file):
    # Criar um novo workbook
    workbook = Workbook()
    # Adicionar uma planilha ativa
    worksheet = workbook.active
    # Salvar o workbook
    workbook.save(output_file)

# Carregar o workbook existente
workbook = load_workbook(output_file)
worksheet = workbook.active

# Itera sobre as linhas do DataFrame
for index, row in df.iterrows():
    text_to_translate = row['esp']
    print(text_to_translate)
    # Verificar se o texto é uma string
    if isinstance(text_to_translate, str):
        if "PERFUME" in text_to_translate.upper():
            colorama.init()
            print(Fore.GREEN + ' _____________________________________________________________________________________')
            print(f"|███████████ Linha {index + 1} contém 'PERFUME': {text_to_translate}")
            print('|███████████__________________________________________________________________________'+Fore.WHITE)
            colorama.deinit()
            row['trad'] = text_to_translate
            worksheet.append(row.tolist())
            continue
        marca_temporario = []
        # Substituir marcas no texto original por um marcador temporário
        for marca in marcas:
            pattern = rf'\b{re.escape(marca)}\b'
            if re.search(pattern, text_to_translate):
                marca_temporario.append(f' {marca} ')
                text_to_translate = re.sub(pattern, '=', text_to_translate)
        translated_text = traduzir(text_to_translate)
        # Restaurar as marcas no texto traduzido
        if marca_temporario != []:
            print(marca_temporario)
            for marca in marca_temporario:
                if '=' in translated_text:
                    translated_text = translated_text.replace('=', marca, 1)
        translated_text = clean_string(translated_text)
        row['trad'] = translated_text
    else:
        # Se não for string, manter o valor original
        row['trad'] = text_to_translate

    worksheet.append(row.tolist())
    
    colorama.init()
    print(Fore.RED + ' _____________________________________________________________________________________')
    print(f"|███████████ Linha {index + 1} traduzida: {row['trad']}")
    print('|███████████__________________________________________________________________________')
    print(Fore.WHITE+'')
    colorama.deinit()
# Salvar o workbook
workbook.save(output_file)
print("Tradução completa!")